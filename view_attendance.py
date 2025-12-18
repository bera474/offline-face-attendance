#!/usr/bin/env python3
"""
Attendance Viewer Script (offline-safe)
- If --date is omitted, shows the most-recent date present in the DB.
- Handles timestamps stored as ISO 'YYYY-MM-DDTHH:MM:SS(.sss)Z' or 'YYYY-MM-DD HH:MM:SS'.
"""

import sqlite3
import csv
from datetime import datetime, timedelta
from pathlib import Path
import sys
import argparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[INFO] openpyxl not installed. Excel export will use CSV fallback.")
    print("       To enable Excel: pip install openpyxl\n")


# ---------- DB path helpers ----------
def default_db_path():
    script_dir = Path(__file__).parent
    return str(script_dir / "attendance.db")


def ensure_db_exists(path):
    if not Path(path).exists():
        print(f"❌ Database not found at {path}")
        sys.exit(1)


# ---------- Date helpers ----------
def latest_date_in_db(db_path: str) -> str | None:
    """
    Return the latest calendar date (YYYY-MM-DD) present in attendance.ts.
    Works for ISO '...T...Z' and ' ' formats.
    """
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT DATE(
            REPLACE(REPLACE(ts, 'T', ' '), 'Z', '')
        ) AS d
        FROM attendance
        WHERE ts IS NOT NULL
        ORDER BY d DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    conn.close()
    return row[0] if row and row[0] else None


def day_bounds(date_str: str) -> tuple[str, str]:
    """Return [start, end) bounds for the given local date."""
    start = f"{date_str} 00:00:00"
    end = (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d 00:00:00")
    return start, end


# ---------- Queries (normalized timestamps) ----------
def get_attendance_data(db_path, date=None, student_name=None):
    """
    Fetch attendance rows for a date window.
    Returns: list of (name, ts, confidence, roll, class).
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    start, end = day_bounds(date)
    ts_norm = "REPLACE(REPLACE(a.ts, 'T', ' '), 'Z', '')"

    if student_name:
        cursor.execute(f"""
            SELECT s.name, a.ts, ROUND(a.confidence, 3) AS confidence,
                   s.roll, s.class
            FROM attendance a
            JOIN students s ON a.student_id = s.id
            WHERE {ts_norm} >= ? AND {ts_norm} < ? AND s.name LIKE ?
            ORDER BY {ts_norm} DESC
        """, (start, end, f"%{student_name}%"))
    else:
        cursor.execute(f"""
            SELECT s.name, a.ts, ROUND(a.confidence, 3) AS confidence,
                   s.roll, s.class
            FROM attendance a
            JOIN students s ON a.student_id = s.id
            WHERE {ts_norm} >= ? AND {ts_norm} < ?
            ORDER BY {ts_norm} DESC
        """, (start, end))

    rows = cursor.fetchall()
    conn.close()
    return rows


def get_summary_data(db_path, date=None):
    """
    Summary per student for a date window.
    Returns: list of (sid, name, roll, class, marks, avg_conf)
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    start, end = day_bounds(date)
    ts_norm = "REPLACE(REPLACE(a.ts, 'T', ' '), 'Z', '')"

    cursor.execute(f"""
        SELECT s.id, s.name, s.roll, s.class,
               COUNT(a.id) AS marks,
               ROUND(AVG(a.confidence), 3) AS avg_confidence
        FROM students s
        LEFT JOIN attendance a
               ON s.id = a.student_id
              AND {ts_norm} >= ?
              AND {ts_norm} < ?
        GROUP BY s.id
        ORDER BY s.name
    """, (start, end))

    rows = cursor.fetchall()
    conn.close()
    return rows


# ---------- Outputs ----------
def create_excel_file(db_path, output_file="attendance_report.xlsx", date=None):
    if not EXCEL_AVAILABLE:
        print("⚠️  openpyxl not installed. Use --csv instead.")
        return False

    data = get_attendance_data(db_path, date)
    summary = get_summary_data(db_path, date)

    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1
    ws1 = wb.create_sheet("Detailed Attendance", 0)
    ws1['A1'] = f"Detailed Attendance - {date}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:E1')

    headers = ['Student Name', 'Roll No', 'Class', 'Time', 'Confidence']
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, (name, ts, conf, roll, cls) in enumerate(data, 4):
        ws1.cell(row=r, column=1).value = name
        ws1.cell(row=r, column=2).value = roll or "-"
        ws1.cell(row=r, column=3).value = cls or "-"
        ws1.cell(row=r, column=4).value = ts
        ws1.cell(row=r, column=5).value = conf
        ws1.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        if r % 2 == 0:
            for col in range(1, 6):
                ws1.cell(row=r, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 26
    ws1.column_dimensions['E'].width = 12

    if data:
        stats_row = len(data) + 5
        ws1[f'A{stats_row}'] = "Total Records:"
        ws1[f'B{stats_row}'] = len(data)
        ws1[f'A{stats_row}'].font = Font(bold=True)

    # Sheet 2
    ws2 = wb.create_sheet("Daily Summary", 1)
    ws2['A1'] = f"Daily Summary - {date}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:F1')

    headers2 = ['Student Name', 'Roll No', 'Class', 'Status', 'Marks', 'Avg Confidence']
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=3, column=i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    present = 0
    for r, (sid, name, roll, cls, marks, avg_conf) in enumerate(summary, 4):
        ws2.cell(row=r, column=1).value = name
        ws2.cell(row=r, column=2).value = roll or "-"
        ws2.cell(row=r, column=3).value = cls or "-"
        status = "✓ Present" if marks and marks > 0 else "✗ Absent"
        if marks and marks > 0: present += 1
        ws2.cell(row=r, column=4).value = status
        ws2.cell(row=r, column=5).value = marks or 0
        ws2.cell(row=r, column=6).value = avg_conf or 0
        for col in [4, 5, 6]:
            ws2.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        if status.startswith("✓"):
            ws2.cell(row=r, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws2.cell(row=r, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        if r % 2 == 0:
            for col in range(1, 7):
                if col != 4:
                    ws2.cell(row=r, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 15

    total = len(summary)
    absent = total - present
    stats_row = total + 5
    ws2[f'A{stats_row}'] = "Summary:"
    ws2[f'A{stats_row}'].font = Font(bold=True)
    ws2[f'A{stats_row + 1}'] = "Total Students:"
    ws2[f'B{stats_row + 1}'] = total
    ws2[f'A{stats_row + 2}'] = "Present:"
    ws2[f'B{stats_row + 2}'] = present
    ws2[f'B{stats_row + 2}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws2[f'A{stats_row + 3}'] = "Absent:"
    ws2[f'B{stats_row + 3}'] = absent
    ws2[f'B{stats_row + 3}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb.save(output_file)
    return True


def create_csv_file(db_path, output_file="attendance_report.csv", date=None):
    data = get_attendance_data(db_path, date)
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    with open(output_file, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Attendance Report', date])
        w.writerow([])
        w.writerow(['Student Name', 'Roll No', 'Class', 'Time', 'Confidence'])
        for name, ts, conf, roll, cls in data:
            w.writerow([name, roll or '-', cls or '-', ts, conf])
        w.writerow([])
        w.writerow(['Total Records:', len(data)])
    return True


def print_table(db_path, date=None):
    data = get_attendance_data(db_path, date)
    summary = get_summary_data(db_path, date)
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    print("\n" + "=" * 100)
    print(f"DETAILED ATTENDANCE REPORT - {date}".center(100))
    print("=" * 100)
    if data:
        print(f"\n{'Student Name':<25} {'Roll':<15} {'Class':<15} {'Time':<30} {'Confidence':<10}")
        print("-" * 100)
        for name, ts, conf, roll, cls in data:
            print(f"{name:<25} {str(roll):<15} {str(cls):<15} {str(ts):<30} {str(conf):<10}")
        print(f"\nTotal Records: {len(data)}")
    else:
        print("\n❌ No attendance records found for this date.\n")

    print("\n" + "=" * 100)
    print(f"DAILY SUMMARY - {date}".center(100))
    print("=" * 100)
    if summary:
        print(f"\n{'Student Name':<25} {'Roll':<15} {'Class':<15} {'Status':<15} {'Marks':<10} {'Avg Confidence':<10}")
        print("-" * 100)
        present = 0
        for sid, name, roll, cls, marks, avg_conf in summary:
            status = "✓ Present" if marks and marks > 0 else "✗ Absent"
            if marks and marks > 0: present += 1
            print(f"{name:<25} {str(roll):<15} {str(cls):<15} {status:<15} {str(marks):<10} {str(avg_conf):<10}")
        total = len(summary)
        absent = total - present
        print("\n" + "-" * 100)
        print(f"Total Students: {total} | Present: {present} | Absent: {absent}")
        print("=" * 100 + "\n")
    else:
        print("\n❌ No students found in database.\n")


def main():
    p = argparse.ArgumentParser(
        description="View attendance records (auto-picks latest date if --date not provided)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python view_attendance.py                    # Show latest day's attendance in terminal
  python view_attendance.py --excel            # Create Excel for latest day
  python view_attendance.py --csv              # Create CSV for latest day
  python view_attendance.py --all              # Excel + CSV + print
  python view_attendance.py --date 2025-12-13  # View specific date
  python view_attendance.py --name "John"      # Filter by student name
  python view_attendance.py --db /path/to/attendance.db
        """
    )
    p.add_argument('--excel', action='store_true', help='Create Excel file')
    p.add_argument('--csv', action='store_true', help='Create CSV file')
    p.add_argument('--all', action='store_true', help='Create both Excel and CSV')
    p.add_argument('--date', help='Date YYYY-MM-DD (default: latest in DB)')
    p.add_argument('--name', help='Filter by student name (Detailed view only)')
    p.add_argument('--output', help='Output filename (without extension)')
    p.add_argument('--db', help='Path to attendance.db (default: alongside this script)')
    args = p.parse_args()

    db_path = args.db or default_db_path()
    ensure_db_exists(db_path)

    # Decide effective date: user-provided or latest in DB
    effective_date = args.date
    if not effective_date:
        latest = latest_date_in_db(db_path)
        if latest:
            effective_date = latest
            print(f"ℹ️  No --date given. Showing latest date in DB: {effective_date}")
        else:
            # No data yet; keep None so the functions print "no records"
            effective_date = None

    # Validate if provided
    if effective_date:
        try:
            datetime.strptime(effective_date, "%Y-%m-%d")
        except ValueError:
            print("❌ Invalid date format. Use YYYY-MM-DD")
            sys.exit(1)

    # If just viewing
    if not args.excel and not args.csv and not args.all:
        print_table(db_path, effective_date if not args.name else None)  # table shows both sections
        # If filtering by name, show only detailed section filtered:
        if args.name:
            data = get_attendance_data(db_path, effective_date, args.name)
            print("\nFiltered by name:\n")
            if data:
                for name, ts, conf, roll, cls in data:
                    print(f"- {name} | {ts} | conf={conf} | roll={roll or '-'} | class={cls or '-'}")
            else:
                print("❌ No matching records.")
        return

    # File outputs
    out_base = args.output or "attendance_report"
    if args.excel or args.all:
        if EXCEL_AVAILABLE:
            if create_excel_file(db_path, f"{out_base}.xlsx", effective_date):
                print(f"✅ Excel file created: {out_base}.xlsx")
        else:
            print("⚠️  openpyxl not installed. Skipping Excel. Install: pip install openpyxl")

    if args.csv or args.all:
        if create_csv_file(db_path, f"{out_base}.csv", effective_date):
            print(f"✅ CSV file created: {out_base}.csv")

    # Always print a quick table at the end
    print_table(db_path, effective_date)


if __name__ == "__main__":
    main()
